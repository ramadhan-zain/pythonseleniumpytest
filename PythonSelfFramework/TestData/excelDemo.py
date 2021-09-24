import openpyxl

book = openpyxl.load_workbook(r"E:\PythonTestingCourse\PythonSelfFramework\TestData\PythonDemo.xlsx")
sheet = book.active

dict_data = {}

cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Ramadhan"

print(sheet.cell(2, 2).value)

print(f"Maximum row: {sheet.max_row} :: Maximum column: {sheet.max_column}")

print(f"Value on A5: {sheet['A5'].value}")

for i in range(1, sheet.max_row + 1):                           # to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):                # to get columns
            print(sheet.cell(row=i, column=j).value)
            dict_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict_data)



