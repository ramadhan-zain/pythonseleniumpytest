import openpyxl


class HomePageData:
    test_HomePage_data = [{"first name": "firstname", "last name": "lastname", "gender": "Male"},
                          {"first name": "firstname1", "last name": "lastname2", "gender": "Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        dict_data = {}

        book = openpyxl.load_workbook(r"E:\PythonTestingCourse\PythonSelfFramework\TestData\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    dict_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict_data]
